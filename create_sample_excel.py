import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Crear datos de ejemplo variados para una tienda
sample_products = [
    {
        'name': 'Coca Cola 600ml',
        'price': 18.50,
        'stock': 24
    },
    {
        'name': 'Sabritas Original 45g',
        'price': 15.00,
        'stock': 30
    },
    {
        'name': 'Gansito',
        'price': 20.00,
        'stock': 15
    },
    {
        'name': 'Agua Bonafont 1L',
        'price': 16.50,
        'stock': 20
    },
    {
        'name': 'Pan Bimbo Grande',
        'price': 45.00,
        'stock': 8
    }
]

# Crear DataFrame
df = pd.DataFrame(sample_products)

# Guardar como Excel
excel_file = 'plantilla_productos.xlsx'
df.to_excel(excel_file, index=False, sheet_name='Productos')

# Abrir el archivo con openpyxl para darle formato
wb = openpyxl.load_workbook(excel_file)
ws = wb['Productos']

# Dar formato al encabezado
header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Ajustar el ancho de las columnas
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

# Agregar hoja de instrucciones
ws_instructions = wb.create_sheet('Instrucciones', 0)
instructions = [
    ['INSTRUCCIONES PARA LLENAR LA PLANTILLA DE PRODUCTOS'],
    [''],
    ['1. La plantilla se encuentra en la hoja "Productos"'],
    ['2. Columnas requeridas:'],
    ['   - name: Nombre del producto (obligatorio)'],
    ['   - price: Precio del producto (obligatorio)'],
    ['   - stock: Cantidad en inventario (opcional, por defecto será 0)'],
    [''],
    ['3. Instrucciones de llenado:'],
    ['   - No cambiar los nombres de las columnas'],
    ['   - Puede agregar tantos productos como necesite'],
    ['   - Los precios deben ser números (ejemplo: 18.50)'],
    ['   - El stock debe ser un número entero'],
    [''],
    ['4. Ejemplo de productos incluidos:'],
    ['   - Puede borrar los productos de ejemplo y agregar los suyos'],
    ['   - O puede agregar sus productos debajo de los ejemplos'],
    [''],
    ['5. Una vez terminado:'],
    ['   - Guarde el archivo'],
    ['   - Súbalo en la sección "Importar Productos desde Excel" de la aplicación'],
]

for row_index, row in enumerate(instructions, 1):
    ws_instructions.cell(row=row_index, column=1, value=row[0])

# Dar formato a las instrucciones
title_font = Font(size=14, bold=True)
ws_instructions['A1'].font = title_font

# Ajustar el ancho de la columna de instrucciones
ws_instructions.column_dimensions['A'].width = 70

wb.save(excel_file)
print(f"Archivo '{excel_file}' creado con éxito.")
print("\nEl archivo contiene dos hojas:")
print("1. 'Instrucciones': Lea esta hoja primero para entender cómo llenar la plantilla")
print("2. 'Productos': Aquí debe ingresar sus productos")
print("\nPuede abrir el archivo con Excel, Numbers o cualquier programa de hojas de cálculo.")
